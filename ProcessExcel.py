# coding: utf-8

import config
import openpyxl
import re
import misc


def getArray(s):
    s = re.split("[,\ ]", s)
    ret_lst = []
    for i in s:
        if i=='':
            continue

        v = i.split('-')
        if len(v)==1:
            ret_lst.append(int(v[0]))
        elif len(v)==2:
            for j in range(int(v[0]), int(v[1])+1):
                ret_lst.append(j)

    return ret_lst


def parseCell(cell_text):
    if cell_text==None or cell_text=="":
        return None

    # preprocess cell texts
    vec = cell_text.split()
    new_vec = []

    # remove teachers' name
    vec = list(filter(lambda x: not(misc.isTeacher(x)), vec))
    for s in vec:
        new_vec.extend(re.split("[\[\]]", s))
    vec = new_vec
    
    # generate class information - name, classroom, week number
    lst = []
    each_class = {}
    for i in vec:
        if len(i)==0:
            pass

        elif misc.isChinese(i[0]):
            if each_class!={}:
                lst.append(each_class)
            each_class = {"name": i}

        elif i[0].isalpha():
            each_class["classroom"] = i

        elif i[0].isdigit():
            if i[-1]=="周":
                new_str = i[0:len(i)-1]
                each_class["weeks"] = getArray(new_str)
    
    if each_class!={}:
        lst.append(each_class)
    return lst


workbook = openpyxl.load_workbook(config.excel_file_path)
worksheet= workbook.active

orig_col = 0
orig_row = 0

try:
    for i in range(1, 10):
        for col in range(1, i+1):
            if worksheet.cell(col, i).value == "星期一":
                orig_col, orig_row = col, i-1
                raise misc.LoopEndError
        for row in range(i, 0, -1):
            if worksheet.cell(i, row).value == "星期一":
                orig_col, orig_row = i, row-1
                raise misc.LoopEndError
except misc.LoopEndError:
    pass


# cal_data (calendar data): array of
#   - Days (0=>Mon., 1=>Tue., ...): array of
#       - time (0=>8:30-10:15, 1=>10:30-12:15, ...): array of
#           - classes (class information): dict containing
#               - classname
#               - classroom
#               - valid week
cal_data = []

for i in range(1, 8):
    every_day = []
    
    for j in range(1, 7):
        parsed = parseCell(worksheet.cell(orig_col+j, orig_row+i).value)
        if parsed!=None:
            every_day.append(parsed)
        else:
            every_day.append([])

    cal_data.append(every_day)

print (cal_data)