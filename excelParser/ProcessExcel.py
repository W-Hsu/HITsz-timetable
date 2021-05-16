# coding=utf-8

from io import BytesIO
from excelParser import lexer, syntaxParser
from interface import config
from errors import ExcelParserError

import warnings
import typing
import openpyxl
import misc

warnings.filterwarnings("ignore")

def process(excel_raw_data) -> typing.Tuple[str, list]:

    if not isinstance(excel_raw_data, bytes):
        raise RuntimeError("")

    workbook = openpyxl.load_workbook(filename=BytesIO(excel_raw_data), read_only=True)
    worksheet= workbook.active

    # Find the origin cell of the sheet
    # Here is a simple illustration of the origin cell:
    # +---------------+--------+--
    # | (Origin Cell) | 星期一 | ...
    # +---------------+--------+--
    # |    第1-2节    | ...
    # +---------------+----
    # | ...           | ...
    orig_col = 3
    orig_row = 1
    try:
        for i in range(1, 10):
            for col in range(1, i+1):
                if worksheet.cell(col, i).value == "星期一":
                    orig_col, orig_row = col, i-1
                    raise misc.StopLoop
            for row in range(i, 0, -1):
                if worksheet.cell(i, row).value == "星期一":
                    orig_col, orig_row = i, row-1
                    raise misc.StopLoop
    except misc.StopLoop:
        pass

    # get calendar name
    cal_name = worksheet.cell(orig_col-2, orig_row).value

    # cal_data (calendar data): array of
    #   - Days (0=>Mon., 1=>Tue., ...): array of
    #       - time (0=>8:30-10:15, 1=>10:30-12:15, ...): array of
    #           - classes (class information): dict containing
    #               - classname
    #               - classroom
    #               - valid week
    cal_data = []

    for i in range(1, 8):
        every_day = []
        for j in range(1, 7):
            try:
                lexer_lst = lexer.parse(worksheet.cell(orig_col+j, orig_row+i).value)
                every_day.append(syntaxParser.parse(lexer_lst))
            except ExcelParserError as ex:
                raise ExcelParserError("Cell" + misc.colRow2ExcelCellName(orig_col+j, orig_row+i) + ": " + str(ex))

        cal_data.append(every_day)
    
    return cal_name, cal_data
